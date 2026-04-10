#!/usr/bin/env python3
"""
yaml-to-sheets.py

Converts a YAML task catalog (MSL-volunteer-opportunities.yaml) into an Excel
spreadsheet (.xlsx) suitable for import into Google Sheets.

One row per task, columns:
  area name, location name, steward, task, task_id, frequency,
  purpose, instructions, supervision, last_date

Validation (aborts on failure):
  - Every task must have a task_id.
  - All task_ids must be unique across the entire file.
  - No duplicate task names within the same location.
"""

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl
import openpyxl.utils
import yaml
from openpyxl.styles import Alignment, Font, PatternFill

COLUMNS = [
    "area name",
    "location name",
    "steward",
    "task",
    "task_id",
    "frequency",
    "purpose",
    "instructions",
    "supervision",
    "last_date",
]

DEFAULT_INPUT = Path("input/MSL-volunteer-opportunities.yaml")
DEFAULT_OUTPUT = Path("input/google-sheet.xlsx")

HEADER_BG = "4472C4"   # blue
HEADER_FG = "FFFFFF"   # white


# ---------------------------------------------------------------------------
# YAML loading / flattening
# ---------------------------------------------------------------------------

def load_yaml(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def detect_shop(data: dict) -> dict:
    for key in ("opportunity", "opportunities", "tasks_list"):
        if key in data:
            return data[key]["shop"]
    raise ValueError(
        "Unknown YAML format — root key must be 'opportunity', 'opportunities', or 'tasks_list'. "
        f"Got: {list(data.keys())}"
    )


def extract_rows(shop: dict) -> list[dict]:
    """Flatten shop hierarchy into one dict per task, preserving order."""
    rows = []
    for area in shop.get("area", []):
        area_name = area.get("name", "")
        for loc in area.get("location", []):
            loc_name = loc.get("name", "")
            # tolerate 'stward' typo present in some legacy YAML
            steward = loc.get("steward", loc.get("stward", ""))
            tasks = loc.get("work_tasks", loc.get("task", [])) or []
            for t in tasks:
                if not isinstance(t, dict):
                    t = {"task": str(t)}
                rows.append({
                    "area name":    area_name,
                    "location name": loc_name,
                    "steward":      steward,
                    "task":         t.get("task", ""),
                    "task_id":      t.get("task_id", ""),
                    "frequency":    t.get("frequency", ""),
                    "purpose":      t.get("purpose", ""),
                    "instructions": t.get("instructions", ""),
                    "supervision":  t.get("supervision", ""),
                    "last_date":    t.get("last_date", ""),
                })
    return rows


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate(rows: list[dict]) -> None:
    errors: list[str] = []

    # 1. task_id must be present on every task
    missing = [
        f"  row {i+1}: area={r['area name']!r}  location={r['location name']!r}  task={r['task']!r}"
        for i, r in enumerate(rows)
        if not str(r.get("task_id", "")).strip()
    ]
    if missing:
        errors.append(
            "task_id missing — add a task_id to each of these tasks:\n" + "\n".join(missing)
        )

    # 2. task_ids must be globally unique
    seen: dict[str, int] = {}
    for i, r in enumerate(rows):
        tid = str(r.get("task_id", "")).strip()
        if not tid:
            continue
        if tid in seen:
            errors.append(
                f"Duplicate task_id {tid!r}: row {seen[tid]+1} "
                f"(area={rows[seen[tid]]['area name']!r}, location={rows[seen[tid]]['location name']!r}) "
                f"and row {i+1} "
                f"(area={r['area name']!r}, location={r['location name']!r})"
            )
        else:
            seen[tid] = i

    # 3. no duplicate task names within a location
    loc_tasks: dict[tuple, list[tuple]] = {}
    for i, r in enumerate(rows):
        key = (r["area name"], r["location name"])
        task_lower = r["task"].strip().lower()
        for j, (prior_lower, prior_row) in enumerate(loc_tasks.get(key, [])):
            if prior_lower == task_lower:
                errors.append(
                    f"Duplicate task in area={r['area name']!r}, location={r['location name']!r}: "
                    f"{r['task']!r} (rows {prior_row+1} and {i+1})"
                )
                break
        loc_tasks.setdefault(key, []).append((task_lower, i))

    if errors:
        print("ERROR: Validation failed. Fix the YAML then re-run.\n", file=sys.stderr)
        for err in errors:
            print(f"  {err}\n", file=sys.stderr)
        sys.exit(1)


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------

def backup_existing(path: Path) -> None:
    """If path exists, copy it to path.bak (or .bak1, .bak2, …)."""
    if not path.exists():
        return
    bak = Path(str(path) + ".bak")
    if not bak.exists():
        shutil.copy2(path, bak)
        print(f"Backup: {bak}")
        return
    n = 1
    while True:
        bak = Path(str(path) + f".bak{n}")
        if not bak.exists():
            shutil.copy2(path, bak)
            print(f"Backup: {bak}")
            return
        n += 1


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volunteer Opportunities"

    header_font  = Font(bold=True, color=HEADER_FG)
    header_fill  = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    center_align = Alignment(horizontal="center", wrap_text=True)

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val = row.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value="" if val is None else str(val))

    # Column widths: max content length, capped at 60
    widths = {col: len(col) for col in COLUMNS}
    for row in rows:
        for col in COLUMNS:
            widths[col] = max(widths[col], min(len(str(row.get(col) or "")), 60))
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths[col_name] + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert YAML task catalog to Excel (.xlsx) for Google Sheets import."
    )
    parser.add_argument(
        "--yaml", default=str(DEFAULT_INPUT),
        help=f"Input YAML file (default: {DEFAULT_INPUT})"
    )
    parser.add_argument(
        "--output", default=str(DEFAULT_OUTPUT),
        help=f"Output Excel file (default: {DEFAULT_OUTPUT})"
    )
    args = parser.parse_args()

    yaml_path = Path(args.yaml)
    out_path  = Path(args.output)

    if not yaml_path.exists():
        print(f"ERROR: YAML not found: {yaml_path}", file=sys.stderr)
        sys.exit(1)

    data  = load_yaml(yaml_path)
    shop  = detect_shop(data)
    rows  = extract_rows(shop)
    validate(rows)
    backup_existing(out_path)
    write_xlsx(rows, out_path)
    print(f"Wrote {len(rows)} rows → {out_path}")


if __name__ == "__main__":
    main()
