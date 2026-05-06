"""tests/test_yaml_to_sheets.py

Tests for scripts/yaml-to-sheets.py: load_yaml, detect_shop, extract_rows,
validate, backup_existing, write_xlsx.
"""
import importlib.util
from pathlib import Path

import pytest

_spec = importlib.util.spec_from_file_location(
    "yaml_to_sheets",
    Path(__file__).parent.parent / "scripts" / "yaml-to-sheets.py",
)
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)
load_yaml = _mod.load_yaml
detect_shop = _mod.detect_shop
extract_rows = _mod.extract_rows
validate = _mod.validate
backup_existing = _mod.backup_existing
write_xlsx = _mod.write_xlsx
COLUMNS = _mod.COLUMNS


SAMPLE_SHOP = {
    "name": "Makersmiths Leesburg",
    "address": "123 Main St",
    "area": [
        {
            "name": "Main Level",
            "location": [
                {
                    "name": "Metalshop",
                    "steward": "Brad Hess",
                    "work_tasks": [
                        {
                            "task": "Wipe machines",
                            "task_id": "MSL-METAL-001",
                            "frequency": "Weekly",
                            "purpose": "Cleanliness",
                            "instructions": "Use rag",
                            "supervision": "None",
                            "last_date": "2024-01-01",
                        },
                        {
                            "task": "Sweep floor",
                            "task_id": "MSL-METAL-002",
                            "frequency": "Daily",
                            "purpose": "Safety",
                            "instructions": "Use broom",
                            "supervision": "None",
                            "last_date": "2024-01-02",
                        },
                    ],
                }
            ],
        }
    ],
}


# ---------------------------------------------------------------------------
# load_yaml
# ---------------------------------------------------------------------------

def test_load_yaml_reads_simple_file(tmp_path) -> None:
    f = tmp_path / "test.yaml"
    f.write_text("key: value\n")
    assert load_yaml(f)["key"] == "value"


def test_load_yaml_reads_nested_structure(tmp_path) -> None:
    f = tmp_path / "test.yaml"
    f.write_text("a:\n  b:\n    c: 42\n")
    assert load_yaml(f)["a"]["b"]["c"] == 42


# ---------------------------------------------------------------------------
# detect_shop
# ---------------------------------------------------------------------------

def test_detect_shop_opportunities_key() -> None:
    result = detect_shop({"opportunities": {"shop": {"name": "Opp"}}})
    assert result["name"] == "Opp"


def test_detect_shop_opportunity_key() -> None:
    result = detect_shop({"opportunity": {"shop": {"name": "Single"}}})
    assert result["name"] == "Single"


def test_detect_shop_tasks_list_key() -> None:
    result = detect_shop({"tasks_list": {"shop": {"name": "List"}}})
    assert result["name"] == "List"


def test_detect_shop_unknown_key_raises_value_error() -> None:
    with pytest.raises(ValueError, match="Unknown YAML format"):
        detect_shop({"bad_key": {}})


# ---------------------------------------------------------------------------
# extract_rows
# ---------------------------------------------------------------------------

def test_extract_rows_returns_correct_count() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    assert len(rows) == 2


def test_extract_rows_area_name() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    assert rows[0]["area name"] == "Main Level"


def test_extract_rows_location_name() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    assert rows[0]["location name"] == "Metalshop"


def test_extract_rows_steward() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    assert rows[0]["steward"] == "Brad Hess"


def test_extract_rows_task_name() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    assert rows[0]["task"] == "Wipe machines"


def test_extract_rows_task_id() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    assert rows[0]["task_id"] == "MSL-METAL-001"


def test_extract_rows_frequency() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    assert rows[0]["frequency"] == "Weekly"


def test_extract_rows_all_columns_present() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    for col in COLUMNS:
        assert col in rows[0], f"Missing column: {col}"


def test_extract_rows_order_preserved() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    assert rows[0]["task"] == "Wipe machines"
    assert rows[1]["task"] == "Sweep floor"


def test_extract_rows_unknown_steward_key_returns_empty() -> None:
    """'stward' is not a recognised key; steward should default to ''."""
    shop = {
        "area": [
            {
                "name": "Area",
                "location": [
                    {
                        "name": "Loc",
                        "stward": "Wrong Key Person",  # misspelled — not read
                        "work_tasks": [{"task": "T", "task_id": "X-001"}],
                    }
                ],
            }
        ]
    }
    rows = extract_rows(shop)
    assert rows[0]["steward"] == ""


def test_extract_rows_plain_string_task() -> None:
    shop = {
        "area": [
            {
                "name": "Area",
                "location": [
                    {
                        "name": "Loc",
                        "steward": "S",
                        "work_tasks": ["plain string task"],
                    }
                ],
            }
        ]
    }
    rows = extract_rows(shop)
    assert rows[0]["task"] == "plain string task"


def test_extract_rows_multiple_areas_flattened() -> None:
    shop = {
        "area": [
            {
                "name": "Area A",
                "location": [
                    {"name": "Loc A", "steward": "S", "work_tasks": [{"task": "T1", "task_id": "A-001"}]},
                ],
            },
            {
                "name": "Area B",
                "location": [
                    {"name": "Loc B", "steward": "S", "work_tasks": [{"task": "T2", "task_id": "B-001"}]},
                ],
            },
        ]
    }
    rows = extract_rows(shop)
    assert len(rows) == 2
    assert rows[0]["area name"] == "Area A"
    assert rows[1]["area name"] == "Area B"


def test_extract_rows_empty_work_tasks_produces_no_rows() -> None:
    shop = {
        "area": [
            {
                "name": "Area",
                "location": [
                    {"name": "Loc", "steward": "S", "work_tasks": []},
                ],
            }
        ]
    }
    rows = extract_rows(shop)
    assert rows == []


# ---------------------------------------------------------------------------
# validate
# ---------------------------------------------------------------------------

def test_validate_passes_on_valid_data() -> None:
    rows = extract_rows(SAMPLE_SHOP)
    validate(rows)  # should not raise


def test_validate_fails_on_missing_task_id() -> None:
    rows = [{"area name": "A", "location name": "L", "task": "T", "task_id": ""}]
    with pytest.raises(SystemExit):
        validate(rows)


def test_validate_fails_on_whitespace_only_task_id() -> None:
    rows = [{"area name": "A", "location name": "L", "task": "T", "task_id": "   "}]
    with pytest.raises(SystemExit):
        validate(rows)


def test_validate_fails_on_duplicate_task_id() -> None:
    rows = [
        {"area name": "A", "location name": "L", "task": "T1", "task_id": "X-001"},
        {"area name": "A", "location name": "L", "task": "T2", "task_id": "X-001"},
    ]
    with pytest.raises(SystemExit):
        validate(rows)


def test_validate_fails_on_duplicate_task_name_in_same_location() -> None:
    rows = [
        {"area name": "A", "location name": "L", "task": "Same Task", "task_id": "X-001"},
        {"area name": "A", "location name": "L", "task": "Same Task", "task_id": "X-002"},
    ]
    with pytest.raises(SystemExit):
        validate(rows)


def test_validate_duplicate_task_name_case_insensitive() -> None:
    rows = [
        {"area name": "A", "location name": "L", "task": "wipe machines", "task_id": "X-001"},
        {"area name": "A", "location name": "L", "task": "Wipe Machines", "task_id": "X-002"},
    ]
    with pytest.raises(SystemExit):
        validate(rows)


def test_validate_allows_same_task_name_in_different_locations() -> None:
    rows = [
        {"area name": "A", "location name": "Loc 1", "task": "Same Task", "task_id": "X-001"},
        {"area name": "A", "location name": "Loc 2", "task": "Same Task", "task_id": "X-002"},
    ]
    validate(rows)  # should not raise


def test_validate_allows_same_task_name_in_different_areas() -> None:
    rows = [
        {"area name": "Area 1", "location name": "L", "task": "Same Task", "task_id": "X-001"},
        {"area name": "Area 2", "location name": "L", "task": "Same Task", "task_id": "X-002"},
    ]
    validate(rows)  # should not raise


# ---------------------------------------------------------------------------
# backup_existing
# ---------------------------------------------------------------------------

def test_backup_creates_bak_file(tmp_path) -> None:
    f = tmp_path / "out.xlsx"
    f.write_bytes(b"data")
    backup_existing(f)
    assert (tmp_path / "out.xlsx.bak").exists()


def test_backup_nonexistent_file_does_nothing(tmp_path) -> None:
    f = tmp_path / "nonexistent.xlsx"
    backup_existing(f)  # should not raise or create files


def test_backup_bak_exists_creates_bak1(tmp_path) -> None:
    f = tmp_path / "out.xlsx"
    f.write_bytes(b"data")
    (tmp_path / "out.xlsx.bak").write_bytes(b"old backup")
    backup_existing(f)
    assert (tmp_path / "out.xlsx.bak1").exists()


def test_backup_preserves_original_file(tmp_path) -> None:
    f = tmp_path / "out.xlsx"
    f.write_bytes(b"original")
    backup_existing(f)
    assert f.read_bytes() == b"original"


# ---------------------------------------------------------------------------
# write_xlsx
# ---------------------------------------------------------------------------

def test_write_xlsx_creates_output_file(tmp_path) -> None:
    rows = extract_rows(SAMPLE_SHOP)
    out = tmp_path / "test.xlsx"
    write_xlsx(rows, out)
    assert out.exists()


def test_write_xlsx_header_row_matches_columns(tmp_path) -> None:
    import openpyxl
    rows = extract_rows(SAMPLE_SHOP)
    out = tmp_path / "test.xlsx"
    write_xlsx(rows, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)]
    assert headers == COLUMNS


def test_write_xlsx_data_row_task_value(tmp_path) -> None:
    import openpyxl
    rows = extract_rows(SAMPLE_SHOP)
    out = tmp_path / "test.xlsx"
    write_xlsx(rows, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    task_col = COLUMNS.index("task") + 1
    assert ws.cell(2, task_col).value == "Wipe machines"


def test_write_xlsx_row_count_matches_data(tmp_path) -> None:
    import openpyxl
    rows = extract_rows(SAMPLE_SHOP)
    out = tmp_path / "test.xlsx"
    write_xlsx(rows, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # Row 1 = header; rows 2..N = data
    assert ws.max_row == len(rows) + 1


def test_write_xlsx_creates_parent_dirs(tmp_path) -> None:
    rows = extract_rows(SAMPLE_SHOP)
    out = tmp_path / "subdir" / "nested" / "test.xlsx"
    write_xlsx(rows, out)
    assert out.exists()
